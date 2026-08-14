"""XLSX 원본을 변경하지 않고 스키마와 기본 데이터 품질을 프로파일링한다."""

from __future__ import annotations

import argparse
import hashlib
import json
import tempfile
from collections import Counter
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from retail_demand_mlops.ingestion.download import (
    DEFAULT_TARGET_PATH,
    calculate_sha256,
)


DEFAULT_REPORT_PATH = Path("data/profiles/online_retail_II.profile.json")


class DatasetProfileError(ValueError):
    """워크북 구조가 프로파일링할 수 없는 상태일 때 발생하는 예외."""


def _is_missing(value: Any) -> bool:
    """빈 문자열도 데이터가 없는 값으로 일관되게 처리한다."""
    return value is None or (isinstance(value, str) and not value.strip())


def _type_name(value: Any) -> str:
    """Python 내부 타입을 보고서에서 읽기 쉬운 안정적인 이름으로 바꾼다."""
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, time):
        return "time"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    if isinstance(value, str):
        return "string"
    return type(value).__name__


def _json_value(value: Any) -> Any:
    """날짜와 시간 값을 JSON에서 손실 없이 읽을 수 있는 ISO 문자열로 변환한다."""
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return value


def _update_range(column_profile: dict[str, Any], value: Any) -> None:
    """수치와 날짜 계열에 대해서만 의미 있는 최솟값과 최댓값을 집계한다."""
    is_numeric = isinstance(value, (int, float)) and not isinstance(value, bool)
    if not is_numeric and not isinstance(value, (date, datetime, time)):
        return

    range_type = "numeric" if is_numeric else _type_name(value)
    if column_profile["range_type"] not in (None, range_type):
        # 서로 비교할 수 없는 타입이 섞이면 오해를 막기 위해 범위 집계를 폐기한다.
        column_profile["range_type"] = "mixed"
        column_profile["minimum"] = None
        column_profile["maximum"] = None
        return
    if column_profile["range_type"] == "mixed":
        return

    column_profile["range_type"] = range_type
    column_profile["minimum"] = (
        value
        if column_profile["minimum"] is None
        else min(column_profile["minimum"], value)
    )
    column_profile["maximum"] = (
        value
        if column_profile["maximum"] is None
        else max(column_profile["maximum"], value)
    )


def _row_signature(row: tuple[Any, ...]) -> bytes:
    """전체 행을 보관하지 않고 중복을 셀 수 있도록 고정 길이 서명을 만든다."""
    digest = hashlib.blake2b(digest_size=16)
    for value in row:
        digest.update(type(value).__name__.encode("utf-8"))
        digest.update(b"\x00")
        digest.update(repr(value).encode("utf-8"))
        digest.update(b"\x1f")
    return digest.digest()


def _profile_sheet(worksheet: Any) -> dict[str, Any]:
    """한 시트를 행 단위로 순회하며 스키마와 품질 지표를 집계한다."""
    rows = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as error:
        raise DatasetProfileError(f"빈 시트는 프로파일링할 수 없습니다: {worksheet.title}") from error

    headers = [str(value).strip() if value is not None else "" for value in raw_headers]
    if any(not header for header in headers):
        raise DatasetProfileError(f"빈 컬럼명이 있습니다: {worksheet.title}")
    if len(headers) != len(set(headers)):
        raise DatasetProfileError(f"중복 컬럼명이 있습니다: {worksheet.title}")

    column_profiles = []
    unique_values: list[set[Any]] = []
    for header in headers:
        column_profiles.append(
            {
                "name": header,
                "null_count": 0,
                "observed_types": Counter(),
                "range_type": None,
                "minimum": None,
                "maximum": None,
            }
        )
        unique_values.append(set())

    header_indexes = {header: index for index, header in enumerate(headers)}
    invoice_index = header_indexes.get("Invoice")
    quantity_index = header_indexes.get("Quantity")
    price_index = header_indexes.get("Price")

    row_count = 0
    duplicate_row_count = 0
    cancellation_row_count = 0
    negative_quantity_count = 0
    negative_price_count = 0
    zero_price_count = 0
    row_signatures: set[bytes] = set()

    for raw_row in rows:
        # 헤더보다 짧은 행도 같은 컬럼 수로 맞춰 누락값이 정확히 집계되게 한다.
        row = tuple(raw_row[index] if index < len(raw_row) else None for index in range(len(headers)))
        row_count += 1

        signature = _row_signature(row)
        if signature in row_signatures:
            duplicate_row_count += 1
        else:
            row_signatures.add(signature)

        for index, value in enumerate(row):
            column_profile = column_profiles[index]
            if _is_missing(value):
                column_profile["null_count"] += 1
                continue
            column_profile["observed_types"][_type_name(value)] += 1
            unique_values[index].add(value)
            _update_range(column_profile, value)

        invoice = row[invoice_index] if invoice_index is not None else None
        if isinstance(invoice, str) and invoice.upper().startswith("C"):
            cancellation_row_count += 1

        quantity = row[quantity_index] if quantity_index is not None else None
        if isinstance(quantity, (int, float)) and quantity < 0:
            negative_quantity_count += 1

        price = row[price_index] if price_index is not None else None
        if isinstance(price, (int, float)):
            negative_price_count += int(price < 0)
            zero_price_count += int(price == 0)

    for index, column_profile in enumerate(column_profiles):
        column_profile["unique_count"] = len(unique_values[index])
        column_profile["observed_types"] = dict(
            sorted(column_profile["observed_types"].items())
        )

        range_type = column_profile["range_type"]
        observed_types = set(column_profile["observed_types"])
        compatible_types = {
            "numeric": {"integer", "number"},
            "date": {"date"},
            "datetime": {"datetime"},
            "time": {"time"},
        }.get(range_type, set())
        if not observed_types.issubset(compatible_types):
            # 식별자처럼 문자열과 숫자가 섞인 컬럼에는 부분적인 숫자 범위를 표시하지 않는다.
            column_profile["range_type"] = None
            column_profile["minimum"] = None
            column_profile["maximum"] = None

        column_profile["minimum"] = _json_value(column_profile["minimum"])
        column_profile["maximum"] = _json_value(column_profile["maximum"])
        if column_profile["range_type"] is None:
            del column_profile["range_type"]
            del column_profile["minimum"]
            del column_profile["maximum"]

    return {
        "name": worksheet.title,
        "row_count": row_count,
        "column_count": len(headers),
        "columns": column_profiles,
        "quality": {
            "duplicate_row_count": duplicate_row_count,
            "cancellation_row_count": cancellation_row_count,
            "negative_quantity_count": negative_quantity_count,
            "negative_price_count": negative_price_count,
            "zero_price_count": zero_price_count,
        },
    }


def profile_workbook(source_path: Path) -> dict[str, Any]:
    """XLSX를 읽기 전용 모드로 한 번 순회하고 프로파일 사전을 반환한다."""
    if not source_path.exists():
        raise DatasetProfileError(f"원본 XLSX가 없습니다: {source_path}")

    # read_only는 셀을 스트리밍하고 data_only는 수식 대신 저장된 결과만 읽는다.
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheets = [_profile_sheet(worksheet) for worksheet in workbook.worksheets]
    finally:
        workbook.close()

    return {
        "source_path": str(source_path),
        "source_sha256": calculate_sha256(source_path),
        "sheet_count": len(sheets),
        "total_row_count": sum(sheet["row_count"] for sheet in sheets),
        "sheets": sheets,
    }


def write_profile(report: dict[str, Any], report_path: Path) -> None:
    """완성된 JSON만 보이도록 임시 파일을 거쳐 보고서를 원자적으로 저장한다."""
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        dir=report_path.parent,
        prefix=f".{report_path.name}.",
        delete=False,
    ) as temporary_file:
        json.dump(report, temporary_file, ensure_ascii=False, indent=2)
        temporary_file.write("\n")
        temporary_path = Path(temporary_file.name)
    temporary_path.replace(report_path)


def main() -> None:
    """명령행에서 기본 원본의 JSON 프로파일을 생성한다."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_TARGET_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_REPORT_PATH)
    arguments = parser.parse_args()

    report = profile_workbook(arguments.source)
    write_profile(report, arguments.output)
    print(f"프로파일 생성 완료: {arguments.output}")


if __name__ == "__main__":
    main()
