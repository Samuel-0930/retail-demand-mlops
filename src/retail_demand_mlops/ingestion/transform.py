"""UCI XLSX 원본을 날짜순 표준 CSV로 스트리밍 변환한다."""

from __future__ import annotations

import argparse
import csv
import json
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from retail_demand_mlops.ingestion.download import (
    DEFAULT_TARGET_PATH,
    calculate_sha256,
)
from retail_demand_mlops.ingestion.normalization import (
    CANONICAL_SALES_COLUMNS,
    TransactionNormalizationError,
    normalize_transaction,
    validate_source_columns,
)


DEFAULT_CSV_PATH = Path("data/processed/online_retail_II.csv")
DEFAULT_MANIFEST_PATH = Path("data/processed/online_retail_II.csv.manifest.json")


class DatasetTransformError(RuntimeError):
    """원본을 일관된 표준 CSV로 변환할 수 없을 때 발생하는 예외."""


@dataclass(frozen=True)
class ConversionResult:
    """변환 실행 여부와 검증 가능한 행 수를 호출자에게 전달한다."""

    created: bool
    row_count: int
    skipped_overlap_count: int


def _read_manifest(manifest_path: Path) -> dict[str, Any] | None:
    """기존 manifest가 없으면 재생성할 수 있도록 None을 반환한다."""
    if not manifest_path.exists():
        return None
    try:
        return json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise DatasetTransformError(f"manifest를 읽을 수 없습니다: {manifest_path}") from error


def _verified_existing_result(
    source_checksum: str,
    target_path: Path,
    manifest_path: Path,
) -> ConversionResult | None:
    """원본과 출력이 기존 manifest와 일치하면 변환 없이 결과를 재사용한다."""
    manifest = _read_manifest(manifest_path)
    if manifest is None or not target_path.exists():
        return None
    if calculate_sha256(target_path) != manifest.get("target_sha256"):
        raise DatasetTransformError(f"기존 CSV의 체크섬이 일치하지 않습니다: {target_path}")
    if manifest.get("source_sha256") != source_checksum:
        return None
    if manifest.get("columns") != list(CANONICAL_SALES_COLUMNS):
        return None
    return ConversionResult(
        created=False,
        row_count=int(manifest["row_count"]),
        skipped_overlap_count=int(manifest["skipped_overlap_count"]),
    )


def convert_workbook_to_csv(
    source_path: Path,
    target_path: Path,
    manifest_path: Path,
) -> ConversionResult:
    """두 시트를 시간순으로 병합하면서 표준 CSV와 manifest를 생성한다.

    Online Retail II의 두 시트는 2010-12-01부터 2010-12-09까지 완전히
    중복된다. 앞 시트를 기준으로 보존하고 다음 시트에서 이전 시트의 마지막
    시각 이하인 행만 건너뛴다. 같은 시트 안의 원래 중복은 원본 사실로 보존한다.
    """
    if not source_path.exists():
        raise DatasetTransformError(f"원본 XLSX가 없습니다: {source_path}")

    source_checksum = calculate_sha256(source_path)
    existing_result = _verified_existing_result(
        source_checksum,
        target_path,
        manifest_path,
    )
    if existing_result is not None:
        return existing_result

    target_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.parent.mkdir(parents=True, exist_ok=True)

    row_count = 0
    skipped_overlap_count = 0
    previous_sheet_max_datetime: datetime | None = None

    with tempfile.TemporaryDirectory(dir=target_path.parent) as temporary_directory:
        temporary_path = Path(temporary_directory)
        temporary_csv_path = temporary_path / target_path.name

        workbook = load_workbook(source_path, read_only=True, data_only=True)
        try:
            with temporary_csv_path.open("w", encoding="utf-8", newline="") as csv_file:
                writer = csv.DictWriter(csv_file, fieldnames=CANONICAL_SALES_COLUMNS)
                writer.writeheader()

                for worksheet in workbook.worksheets:
                    rows = worksheet.iter_rows(values_only=True)
                    try:
                        headers = next(rows)
                    except StopIteration as error:
                        raise DatasetTransformError(
                            f"빈 시트는 변환할 수 없습니다: {worksheet.title}"
                        ) from error
                    validate_source_columns(headers)

                    current_sheet_max_datetime: datetime | None = None
                    previous_row_datetime: datetime | None = None
                    for source_row_number, values in enumerate(rows, start=2):
                        source_row = dict(zip(headers, values, strict=True))
                        invoice_datetime = source_row.get("InvoiceDate")
                        if not isinstance(invoice_datetime, datetime):
                            raise DatasetTransformError(
                                "거래 시각이 datetime이 아닙니다: "
                                f"sheet={worksheet.title}, row={source_row_number}"
                            )
                        if (
                            previous_row_datetime is not None
                            and invoice_datetime < previous_row_datetime
                        ):
                            raise DatasetTransformError(
                                "시트가 거래 시각 순서로 정렬되어 있지 않습니다: "
                                f"sheet={worksheet.title}, row={source_row_number}"
                            )
                        previous_row_datetime = invoice_datetime
                        current_sheet_max_datetime = invoice_datetime

                        if (
                            previous_sheet_max_datetime is not None
                            and invoice_datetime <= previous_sheet_max_datetime
                        ):
                            skipped_overlap_count += 1
                            continue

                        try:
                            writer.writerow(normalize_transaction(source_row))
                        except TransactionNormalizationError as error:
                            raise DatasetTransformError(
                                "원본 행 정규화에 실패했습니다: "
                                f"sheet={worksheet.title}, row={source_row_number}"
                            ) from error
                        row_count += 1

                    if current_sheet_max_datetime is None:
                        raise DatasetTransformError(
                            f"데이터 행이 없는 시트입니다: {worksheet.title}"
                        )
                    previous_sheet_max_datetime = max(
                        previous_sheet_max_datetime or current_sheet_max_datetime,
                        current_sheet_max_datetime,
                    )
        finally:
            workbook.close()

        target_checksum = calculate_sha256(temporary_csv_path)
        manifest = {
            "source_path": str(source_path),
            "source_sha256": source_checksum,
            "target_path": str(target_path),
            "target_sha256": target_checksum,
            "columns": list(CANONICAL_SALES_COLUMNS),
            "row_count": row_count,
            "skipped_overlap_count": skipped_overlap_count,
        }
        temporary_manifest_path = temporary_path / manifest_path.name
        temporary_manifest_path.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

        temporary_csv_path.replace(target_path)
        temporary_manifest_path.replace(manifest_path)

    return ConversionResult(
        created=True,
        row_count=row_count,
        skipped_overlap_count=skipped_overlap_count,
    )


def main() -> None:
    """명령행에서 기본 UCI 원본을 표준 CSV로 변환한다."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_TARGET_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_CSV_PATH)
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST_PATH)
    arguments = parser.parse_args()

    result = convert_workbook_to_csv(
        arguments.source,
        arguments.output,
        arguments.manifest,
    )
    status = "변환 완료" if result.created else "기존 CSV 검증 완료"
    print(
        f"{status}: rows={result.row_count}, "
        f"skipped_overlap={result.skipped_overlap_count}, path={arguments.output}"
    )


if __name__ == "__main__":
    main()
